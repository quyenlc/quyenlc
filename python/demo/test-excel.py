from zabbix_api import ZabbixAPI
from pprint import pprint
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "DB performce report"
zapi = ZabbixAPI(server="http://mon.marvel.denagames-asia.com/zabbix")
username = 'infra'
password = 'Punch123!'

# Header
ws['A1'] = 'Time'
ws['B1'] = '02/04-07/04'
ws['B2'] = 'Max CPU Ulti (%)'
ws['C2'] = 'Avg CPU Ulti (%)'
ws['D2'] = 'Query Per Second'
ws['E2'] = 'Current Instance Type'
ws['F2'] = 'Max Mem Usage (GB)'
ws['G2'] = 'Avg Mem Usage (GB)'
ws['H2'] = 'Max IOPS Data Volume'
ws['I2'] = 'Avg IOPS Data Volume'
ws['J2'] = 'Size Data Volume'
ws['K2'] = 'Usage Data Volume'
ws['L2'] = 'IOPS'
ws['M2'] = 'Max IOPS Log Volume'
ws['N2'] = 'Avg IOPS Log Volume'
ws['O2'] = 'Size Log Volume'
ws['P2'] = 'Usage Log Volume'
ws['Q2'] = 'IOPS'
ws['R2'] = 'Max Traffic Input'
ws['S2'] = 'Avg Traffic Input'
ws['T2'] = 'Max Traffic Output'
ws['U2'] = 'Avg Traffic Output'
ws['V2'] = 'Note'

#Body


#Save
wb.save('/Users/quyen.le/temp/balances.xlsx')

#Body
